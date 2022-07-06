from odoo import api, fields, models, _
from odoo.exceptions import Warning
import base64
import csv
from io import StringIO, BytesIO
from csv import DictWriter, DictReader
from openpyxl import load_workbook, Workbook
import xlrd
import xlwt


class ImportStockInvnetoryAdjustments(models.TransientModel):
    _name = "ywt.import.stock.inventory.adjustments"
    
    is_validate_inventory = fields.Boolean(string="Validate Inventory")
    
    select_file = fields.Binary(string='Select File')
    datas = fields.Binary(string='File Data')
    
    selectedfile_name = fields.Char(string='Filename', size=512)
    
    import_productby = fields.Selection([('name', 'Name'), ('default_code', 'Internal Reference'), ('barcode', 'Barcode')], default='default_code', string="Import Product By")
    import_filetype = fields.Selection([('csv', 'CSV'), ('xlsx', 'Xlsx')], default="xlsx", string="Import Report Type")
    file_delimiter = fields.Selection([(',', 'Comma')], default=',', string="Delimiter", help="Select a delimiter to process CSV file.")
    
    stocklocation_id = fields.Many2one("stock.location", string="Source Location")
    
    def process_inventory_adjustments(self):
        stock_inventory_obj = self.env['stock.inventory']
        stock_inventory_line_obj = self.env['stock.inventory.line']
        product_product_obj = self.env['product.product']
        inventory_log_obj = self.env['ywt.import.stock.inventory.log']
        
        select_inventory_file = self.validate_process()[0]
        
        if select_inventory_file:
            
            file_name = self.selectedfile_name
            index = file_name.rfind('.')
            flag = 0
            if index == -1:
                flag = 1
            extension = file_name[index + 1:]
                    
            if flag or extension not in ['xlsx', 'csv']:
                raise Warning('Please Provide only .xlsx , .csv file')

            column_header = {}
            
            selected_file = self.select_file
           
            stock_location = self.stocklocation_id
            
            stock_inventory_id = stock_inventory_obj.search([('state', 'in', ['cancel', 'confirm']), ('location_id', '=', stock_location.id)], limit=1)
            if stock_inventory_id:
                raise Warning ("There is already One Inventory %s and it location %s in In Progress State Once you Validate Inventory then After Process Other") % (stock_inventory_id.name, stock_location.display_name)
             
            if not stock_inventory_id:
                vals = self.prepare_invnetory_adjustments_vals(stock_location)
                stock_inventory_id = stock_inventory_obj.create(vals)
             
            log_record = inventory_log_obj.log_record(operation_type="import")
            product_id = False 
            
            if self.import_filetype == 'csv':
                
                self.write({'datas':self.select_file})
                self._cr.commit()
                
                imp_file = BytesIO(base64.decodestring(self.datas))
                csvf = StringIO(imp_file.read().decode())
                file_lines = csv.DictReader(csvf, delimiter=',')        
                for file_line in file_lines:
                    if self.import_productby == 'name':
                        if not file_line.get('Name', {}):
                            continue
                        product_name = file_line.get('Name').strip() 
                        if product_name != None:
                            product_id = product_product_obj.search([('name', '=', product_name), ('type', '=', 'product')], limit=1)
                            if not product_id:
                                msg = "Product Not Found Product Name %s" % (product_name)
                                log_record.post_log_line(msg, type='mismatch')
                    elif  self.import_productby == 'default_code' :
                        if not file_line.get('Default Code'):
                            continue
                        product_code = file_line.get('Default Code').strip()  
                        if product_code != None:
                            product_id = product_product_obj.search([('default_code', '=', product_code), ('type', '=', 'product')], limit=1)
                            if not product_id:
                                msg = "Product Not Found Product Internal Reference %s" % (product_code)
                                log_record.post_log_line(msg, type='mismatch')
                    elif self.import_productby == 'barcode':
                        if not file_line.get('Barcode', {}):
                            continue
                        product_barcode = file_line.get('Barcode').strip() 
                        if product_barcode != None:
                            product_id = product_product_obj.search([('barcode', '=', product_barcode), ('type', '=', 'product')], limit=1)
                            if not product_id:
                                msg = "Product Not Found Product Barcode %s" % (product_barcode)
                                log_record.post_log_line(msg, type='mismatch')
                            
                    quantity = file_line.get('Quantity').strip()
                    if not quantity:
                        raise Warning('Please Enter Quantity Column')
                    if product_id and quantity:
                        line_id = stock_inventory_line_obj.search([('product_id', '=', product_id.id), ('inventory_location_id', '=', stock_location.id), ('inventory_id', '=', stock_inventory_id.id)], limit=1)
                        if line_id:
                            line_id.update({'product_qty':quantity})
                        else:
                            stock_inventory_line_vals = self.prepare_invnetory_adjustments_line_vals(product_id, quantity, stock_location, stock_inventory_id)
                            stock_inventory_line_obj.create(stock_inventory_line_vals)
                            
            if self.import_filetype == 'xlsx':
                column_header = {}
                try:
                    worksheet = self.read_file(selected_file)
                    column_header = self.read_header(worksheet)
                    column_header_value = column_header.values()
                except Exception as e:
                    error_value = str(e)
                    raise Warning(error_value)
                
                file_stock_inventory_lst = self.prepare_list_dict_data(worksheet, column_header)
                
                if self.check_header_validation(column_header_value):
                    for file_line in file_stock_inventory_lst:
                        if self.import_productby == 'name':
                            product_name = file_line.get('name')
                            product_id = product_product_obj.search([('name', '=', product_name), ('type', '=', 'product')], limit=1)
                            if not product_id:
                                msg = "Product Not Found Product Name %s" % (product_name)
                                log_record.post_log_line(msg, type='mismatch')
                        elif self.import_productby == 'default_code':
                            product_code = file_line.get('default code')
                            product_id = product_product_obj.search([('default_code', '=', product_code), ('type', '=', 'product')], limit=1)
                            if not product_id:
                                msg = "Product Not Found Product Internal Reference %s" % (product_code)
                                log_record.post_log_line(msg, type='mismatch')
                            
                        elif self.import_productby == 'barcode':
                            product_barcode = file_line.get('barcode')
                            product_id = product_product_obj.search([('barcode', '=', product_barcode), ('type', '=', 'product')], limit=1)
                            if not product_id:
                                msg = "Product Not Found Product Barcode %s" % (product_barcode)
                                log_record.post_log_line(msg, type='mismatch')
                                               
                        quantity = file_line.get('quantity')
                        if quantity and product_id:
                            line_id = stock_inventory_line_obj.search([('product_id', '=', product_id.id), ('inventory_location_id', '=', stock_location.id), ('inventory_id', '=', stock_inventory_id.id)], limit=1)
                            if line_id:
                                line_id.update({'product_qty':quantity})
                            else:
                                stock_inventory_line_vals = self.prepare_invnetory_adjustments_line_vals(product_id, quantity, stock_location, stock_inventory_id)
                                stock_inventory_line_obj.create(stock_inventory_line_vals)
                                
            if not stock_inventory_id.line_ids:
                stock_inventory_id.sudo().unlink()
                raise Warning('Inventory not created due to some error please check the Import Stock Inventory Log...')
                
            if stock_inventory_id.line_ids and self.is_validate_inventory : 
                stock_inventory_id.action_validate()
                

            if not log_record.log_line_ids:
                log_record.sudo().unlink()

    def validate_process(self):
        if not self.select_file:
            raise Warning('Please Select File to Process...')
        
        return True
    

    def prepare_invnetory_adjustments_vals(self, stock_location):
        sequence_id = self.env.ref('ywt_import_stock_inventory.stock_inventory_seq').ids
        if sequence_id:
            record_name = self.env['ir.sequence'].get_id(sequence_id[0])
        vals = {
            'name': record_name,
            'location_id':stock_location and stock_location.id,
            'company_id' : stock_location and stock_location.company_id.id or False,
            'filter':'partial',
        }
        return vals
    
    def prepare_invnetory_adjustments_line_vals(self, product_id, quantity, stock_location, stock_inventory_id):
        vals = {
            'inventory_id':stock_inventory_id and stock_inventory_id.id,
            'location_id':stock_location and stock_location.id,
            'company_id':stock_location and stock_location.company_id.id,
            'product_id':product_id and product_id.id,
            'product_uom_id':product_id and product_id.uom_id.id,
            'product_qty':quantity,
            }
        return vals
    
    def read_file(self, selected_file):
        try:
            imp_file = BytesIO(base64.decodestring(selected_file))
            workbook = load_workbook(filename=imp_file)
            worksheet = workbook.get_active_sheet()
        except Exception as e:
            error_value = str(e)
            raise Warning(error_value)
        
        return worksheet
    
    def read_header(self, worksheet):
        try:
            column_header = {}
            columns = worksheet.max_column
            columns = columns + 1    
            for row in worksheet.iter_rows(): 
                if row[0].row > 1 or not next((r for r in row if r.row), None):
                    break
                for col in range(1, columns):
                    column_header.update({col : str(worksheet.cell(row=1, column=col).value).lower()})
        except Exception as e:
            error_value = str(e)
            raise Warning(error_value)
        
        return column_header
    
    def check_header_validation(self, column_header_value):
        if self.import_productby == 'name':
            majorities_field = ['name', 'quantity']
        elif self.import_productby == 'default_code':
            majorities_field = ['default code', 'quantity']
        elif self.import_productby == 'barcode':
            majorities_field = ['barcode', 'quantity']
        missing = []
        for field in majorities_field:
            if field not in column_header_value:
                missing.append(field)
            
        if len(missing) > 0:
            raise Warning('Please Add Mandatory Field  First %s in File ' % (missing))
        
        return True
    
    def prepare_list_dict_data(self, worksheet, column_header):
        try:
            data = []
            columns = worksheet.max_column
            columns = columns + 1
            for row_num, row in enumerate(worksheet.iter_rows()):
                row_num += 1
                if row[0].row == 1 or not next((r for r in row if r.row), None):
                    continue
                test_data = {}
                for col in range(1, columns):
                    test_data.update({column_header.get(col) : worksheet.cell(row=row_num, column=col).value})
                data.append(test_data)
        except Exception as e:
            error_value = str(e)
            raise Warning(error_value)
            
        return data   
